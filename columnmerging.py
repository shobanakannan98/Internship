import pandas as pd
from openpyxl import load_workbook

def drop_month_column(file_path):
    print(f"Attempting to open file: {file_path}")
    # Load the Excel file
    xls = pd.ExcelFile(file_path)

    # Extract the FactFinance sheet into a dataframe
    df_fact_finance = pd.read_excel(xls, 'FactFinance')

    # Ensure the column names are stripped of spaces (in case of unexpected white spaces)
    df_fact_finance.columns = df_fact_finance.columns.str.strip()

    # Convert the Month column to Date column in the format dd.mm.yyyy
    df_fact_finance['Date'] = pd.to_datetime(df_fact_finance['Month'].astype(str) + '01', format='%Y%m%d', errors='coerce').dt.strftime('%d.%m.%Y')

    # Drop the Month column
    df_fact_finance.drop(columns=['Month'], inplace=True)

    # Save the transformed data back to the same sheet in the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_fact_finance.to_excel(writer, sheet_name='FactFinance', index=False)

    print("Month column dropped and data saved successfully.")

def transform_fact_finance(file_path):
    print(f"Attempting to open file: {file_path}")
    # Load the Excel file
    xls = pd.ExcelFile(file_path)

    # Extract the FactFinance sheet into a dataframe
    df_fact_finance = pd.read_excel(xls, 'FactFinance')

    # Ensure the column names are stripped of spaces (in case of unexpected white spaces)
    df_fact_finance.columns = df_fact_finance.columns.str.strip()

    # Convert the Month column to Date column in the format dd.mm.yyyy
    df_fact_finance['Date'] = pd.to_datetime(df_fact_finance['Month'].astype(str) + '01', format='%Y%m%d', errors='coerce').dt.strftime('%d.%m.%Y')

    # Drop the Month column
    df_fact_finance.drop(columns=['Month'], inplace=True)

    # Transform the columns in FactFinance
    df_actual = df_fact_finance.copy()
    df_actual['Scenario'] = 'Actual'
    df_actual['Margin'] = df_actual['Margin, actual']
    df_actual['Order Quantity'] = df_actual['Order Quantity, actual']
    df_actual['Sales Amount'] = df_actual['Sales Amount, actual']
    df_actual['Total Product Cost'] = df_actual['Total Product Cost, actual']
    df_actual['Fixed Costs'] = df_actual['Fixed Costs, actual']

    df_budget = df_fact_finance.copy()
    df_budget['Scenario'] = 'Budget'
    df_budget['Margin'] = df_budget['Margin, budget']
    df_budget['Order Quantity'] = df_budget['Order Quantity, budget']
    df_budget['Sales Amount'] = df_budget['Sales Amount, budget']
    df_budget['Total Product Cost'] = df_budget['Total Product Cost, budget']
    df_budget['Fixed Costs'] = df_budget['Fixed Costs, budget']

    # Concatenate the transformed data
    df_transformed = pd.concat([df_actual, df_budget], ignore_index=True)

    # Drop the original margin columns and other actual/budget columns
    df_transformed.drop(columns=[
        'Margin, actual', 'Margin, budget',
        'Order Quantity, actual', 'Order Quantity, budget',
        'Sales Amount, actual', 'Sales Amount, budget',
        'Total Product Cost, actual', 'Total Product Cost, budget',
        'Fixed Costs, actual', 'Fixed Costs, budget'
    ], inplace=True)

    # Save the transformed data back to the same sheet in the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_transformed.to_excel(writer, sheet_name='FactFinance', index=False)

    print("Transformation is successfully saved.")

def create_dim_address(file_path):
    print(f"Attempting to open file: {file_path}")
    # Load the Excel file
    xls = pd.ExcelFile(file_path)

    # Extract the DimCustomer sheet into a dataframe
    df_dim_customer = pd.read_excel(xls, 'DimCustomer')

    # Ensure the column names are stripped of spaces (in case of unexpected white spaces)
    df_dim_customer.columns = df_dim_customer.columns.str.strip()

    # Extract columns 'State', 'Region', 'Country' to create DimAddress
    df_dim_address = df_dim_customer[['State', 'Region', 'Country']].copy()

    # Create a primary key column 'StateID' in DimAddress
    df_dim_address['StateID'] = range(1, len(df_dim_address) + 1)

    # Add 'StateID' as a foreign key in DimCustomer
    df_dim_customer = df_dim_customer.merge(df_dim_address[['State', 'StateID']], on='State', how='left')

    # Save the transformed data back to the same Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_dim_customer.to_excel(writer, sheet_name='DimCustomer', index=False)
        df_dim_address.to_excel(writer, sheet_name='DimAddress', index=False)

    print("DimAddress sheet created and data saved successfully.")

def delete_sheets_and_columns(file_path):
    print(f"Attempting to open file: {file_path}")
    # Load the Excel file
    workbook = load_workbook(file_path)

    # Delete the sheets 'POC 1' and 'POC 2' if they exist
    for sheet_name in ['POC 1', 'POC 2']:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
            print(f"Deleted sheet: {sheet_name}")

    # Save the workbook after deleting the sheets
    workbook.save(file_path)

    # Load the Excel file again to work with pandas
    xls = pd.ExcelFile(file_path)

    # Extract the DimCustomer sheet into a dataframe
    df_dim_customer = pd.read_excel(xls, 'DimCustomer')

    # Ensure the column names are stripped of spaces (in case of unexpected white spaces)
    df_dim_customer.columns = df_dim_customer.columns.str.strip()

    # Delete the columns 'State', 'Region', 'Country' from DimCustomer
    df_dim_customer.drop(columns=['State', 'Region', 'Country'], inplace=True)

    # Save the transformed data back to the same Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_dim_customer.to_excel(writer, sheet_name='DimCustomer', index=False)

    print("Sheets 'POC 1' and 'POC 2' deleted, and columns 'State', 'Region', 'Country' removed from DimCustomer.")

if __name__ == "__main__":
    file_path = r"C:\Users\perumalm\Desktop\GIT\MySQL server\GitHub_Copliot_demo\Excelfiles\Git_Copilot_POC.xlsx"
    # Call the drop_month_column function
    drop_month_column(file_path)

    # Call the transform_fact_finance function
    transform_fact_finance(file_path)

    # Call the create_dim_address function
    create_dim_address(file_path)

    # Call the delete_sheets_and_columns function
    delete_sheets_and_columns(file_path)

