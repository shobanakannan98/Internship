import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
file_path = "C:/Users/perumalm/Desktop/GIT/MySQL server/GitHub_Copliot_demo/Excelfiles/Git_Copilot_POC_1.xlsx"
xls = pd.ExcelFile(file_path)

# Load the sheets into DataFrames
dim_customer = pd.read_excel(xls, 'DimCustomer')
dim_product = pd.read_excel(xls, 'DimProduct')
fact_finance = pd.read_excel(xls, 'FactFinance')

# Create DimAddress DataFrame with unique values from DimCustomer
dim_address = dim_customer[['State', 'Region', 'Country']].drop_duplicates().reset_index(drop=True)

# Add a unique StateID column
dim_address['StateID'] = range(1, len(dim_address) + 1)

# Merge StateID into DimCustomer as a foreign key
dim_customer = dim_customer.merge(dim_address[['State', 'StateID']], on='State', how='left')

# Drop the 'State', 'Region', 'Country' columns in DimCustomer
dim_customer.drop(columns=['State', 'Region', 'Country'], inplace=True)

# Create DimDate DataFrame with unique dates from FactFinance
dim_date = fact_finance[['Date']].drop_duplicates().reset_index(drop=True)

# Add a unique DateID column
dim_date['DateID'] = range(1, len(dim_date) + 1)

# Merge DateID into FactFinance as a foreign key
fact_finance = fact_finance.merge(dim_date[['Date', 'DateID']], on='Date', how='left')

# Drop the 'Date' column in FactFinance
fact_finance.drop(columns=['Date'], inplace=True)

# Load the workbook and remove the specified sheets
workbook = load_workbook(file_path)
if 'POC 1' in workbook.sheetnames:
    del workbook['POC 1']
if 'POC 2' in workbook.sheetnames:
    del workbook['POC 2']

# Save the transformed DataFrames back to the Excel file
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    dim_customer.to_excel(writer, sheet_name='DimCustomer', index=False)
    dim_product.to_excel(writer, sheet_name='DimProduct', index=False)
    fact_finance.to_excel(writer, sheet_name='FactFinance', index=False)
    dim_address.to_excel(writer, sheet_name='DimAddress', index=False)
    dim_date.to_excel(writer, sheet_name='DimDate', index=False)

# Save the workbook after removing the sheets
workbook.save(file_path)

print(f"Transformed data has been saved to {file_path}")